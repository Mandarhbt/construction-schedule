from dotenv import load_dotenv
load_dotenv()

try:
    from ai_planner import generate_activities, test_connection
except Exception:
    def test_connection():
        return "AI planner not available"
    def generate_activities(*args, **kwargs):
        raise RuntimeError("ai_planner not available. Ensure ai_planner.py exists and openai is installed.")

from cpm import calculate_cpm
from calendar_utils import add_working_days, next_working_day
import streamlit as st
import pandas as pd
import json
import os
import re
from collections import defaultdict, deque


# ── WBS Hierarchical Table Renderer ───────────────────────────────────────────
def render_wbs_table(df, data_cols, wbs_col="WBS"):
    """Return an HTML string rendering df as a table with WBS group header rows.

    WBS values are split on ' > ' to build a hierarchy.  Each WBS level gets
    its own header row in a progressively lighter shade of grey.  Activity rows
    are indented proportionally to their WBS depth.

    data_cols : list of column names shown *beside* Activity (WBS is the grouping
                column, shown only in header rows, not as a body column).
    """
    # Progressively lighter grey shades; (CSS, indent-px)
    LEVEL_STYLES = [
        ("background:#455a64;color:#fff;font-weight:700;",   0),   # depth 0
        ("background:#607d8b;color:#fff;font-weight:700;",  12),   # depth 1
        ("background:#90a4ae;color:#263238;font-weight:600;", 24),  # depth 2
        ("background:#b0bec5;color:#263238;font-weight:600;", 36),  # depth 3
        ("background:#cfd8dc;color:#263238;font-weight:500;", 48),  # depth 4+
    ]

    def parse_wbs(val):
        s = str(val).strip() if val is not None else ""
        if s in ("", "nan", "None", "General"):
            return ["General"]
        parts = [p.strip() for p in re.split(r"\s*>\s*", s) if p.strip()]
        return parts if parts else ["General"]

    th = "background:#263238;color:#fff;padding:7px 10px;text-align:left;" \
         "font-size:12px;white-space:nowrap;border-bottom:2px solid #546e7a;"
    total_cols = len(data_cols) + 1  # Activity col + data cols

    html = [
        "<div style='overflow-x:auto;'>",
        "<table style='border-collapse:collapse;width:100%;font-size:12px;"
        "font-family:\"Segoe UI\",Arial,sans-serif;'>",
        "<thead><tr>",
        f"<th style='{th}'>Activity</th>",
    ]
    for col in data_cols:
        html.append(f"<th style='{th}'>{col}</th>")
    html.append("</tr></thead><tbody>")

    current_path = []
    for row_num, (_, row) in enumerate(df.iterrows()):
        wbs_path = parse_wbs(row.get(wbs_col, ""))

        # How many leading levels match current_path?
        common = sum(1 for a, b in zip(current_path, wbs_path) if a == b)
        # Emit WBS header rows for newly entered levels
        for lvl in range(common, len(wbs_path)):
            sty, base_px = LEVEL_STYLES[min(lvl, len(LEVEL_STYLES) - 1)]
            label = wbs_path[lvl]
            pad_left = 10 + base_px
            nbsp_indent = "\u00a0" * (lvl * 3)
            html.append(
                f"<tr><td colspan='{total_cols}' style='{sty}"
                f"padding:5px 10px 5px {pad_left}px;"
                f"border-bottom:1px solid rgba(0,0,0,.18);letter-spacing:.3px;'>"
                f"{nbsp_indent}{label}</td></tr>"
            )
        current_path = list(wbs_path)

        # Activity row
        depth     = len(wbs_path)
        pad_act   = 10 + depth * 16
        row_bg    = "#ffffff" if row_num % 2 == 0 else "#f5f7f9"
        is_crit   = str(row.get("Critical", "")).strip().lower() in ("true", "1", "yes")
        if is_crit:
            row_bg = "#fff8e1"

        base_cell = "padding:4px 10px;border-bottom:1px solid #e0e0e0;white-space:nowrap;"
        crit_text = "color:#c62828;font-weight:600;"

        html.append(f"<tr style='background:{row_bg};'>")
        act_sty = f"padding:4px 10px 4px {pad_act}px;border-bottom:1px solid #e0e0e0;"
        if is_crit:
            act_sty += crit_text
        act_val = str(row.get("Activity", ""))
        nbsp_act = "&nbsp;" * (depth * 2)
        html.append(f"<td style='{act_sty}'>{nbsp_act}{act_val}</td>")

        for col in data_cols:
            raw = row.get(col, "")
            display = "" if (raw is None or str(raw) in ("nan", "None")) else str(raw)
            csty = base_cell
            if col == "Critical":
                if is_crit:
                    display = "\u26d4 Yes"
                    csty += crit_text
                else:
                    display = ""
            elif col == "Float":
                try:
                    if display != "" and int(float(display)) == 0:
                        csty += crit_text
                        display = "0"
                except Exception:
                    pass
            html.append(f"<td style='{csty}'>{display}</td>")
        html.append("</tr>")

    html.append("</tbody></table></div>")
    return "".join(html)


st.set_page_config(page_title="Construction Scheduler", layout="wide", initial_sidebar_state="collapsed")

# ── Analytics ─────────────────────────────────────────────────────────────────
import streamlit.components.v1 as components
components.html(
    """
    <!-- Microsoft Clarity -->
    <script type="text/javascript">
        (function(c,l,a,r,i,t,y){
            c[a]=c[a]||function(){(c[a].q=c[a].q||[]).push(arguments)};
            t=l.createElement(r); t.async=1;
            t.src="https://www.clarity.ms/tag/"+i;
            y=l.getElementsByTagName(r)[0];
            y.parentNode.insertBefore(t,y);
        })(window.parent, window.parent.document, "clarity", "script", "xh5u3qsr2v");
    </script>

    <!-- Google tag (gtag.js) -->
    <script>
        (function() {
            var s = window.parent.document.createElement('script');
            s.async = true;
            s.src = 'https://www.googletagmanager.com/gtag/js?id=G-HDT4HM4ZHS';
            window.parent.document.head.appendChild(s);

            window.parent.dataLayer = window.parent.dataLayer || [];
            window.parent.gtag = function(){ window.parent.dataLayer.push(arguments); };
            window.parent.gtag('js', new Date());
            window.parent.gtag('config', 'G-HDT4HM4ZHS');
        })();
    </script>
    """,
    height=0,
    width=0,
)

st.title("Construction Scheduler")

# ── Sidebar / Hamburger Menu ───────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Construction Scheduler")
    st.caption("**v0.1.0** · 5-Jul-2026")
    st.divider()

    with st.expander("ℹ️ About", expanded=False):
        st.markdown(
            """
            **Construction Scheduler** is an AI-powered project scheduling tool
            built for construction professionals.

            Generate detailed, WBS-structured activity lists using AI or
            built-in templates, compute critical-path dates automatically, and
            export to **Excel** or **Primavera P6 XML** — all in a single
            browser-based workflow.

            **Developed by:** Mandar  
            **Version:** v0.1.0  
            **Release date:** 5-Jul-2026
            """
        )

    with st.expander("✨ Features — v0.1.0", expanded=False):
        st.markdown(
            """
            - 🤖 AI-powered activity generation (GPT-4o)
            - 📋 Template-based schedules (Residential / Commercial / Hospital)
            - 🗂️ WBS hierarchical display & editing
            - 📅 Custom calendar management (working days & holidays)
            - ⚙️ Critical Path Method (CPM) — ES, EF, LS, LF, Float
            - 🏁 Auto-generated project milestones
            - 📝 Scheduling rules editor
            - 💡 Per-activity AI rationale / notes
            - 📥 Styled Excel export (WBS hierarchy mirrored)
            - 📥 Primavera P6 XML export
            - 🔍 Schedule quality log (9 health checks)
            """
        )

    with st.expander("🚀 Upcoming Features", expanded=False):
        st.markdown(
            """
            - 👷 Resource loading on activities
            """
        )

    st.divider()
    st.link_button("💬 Give Feedback", "https://forms.gle/S1gmY7YysDDUMnwt5",
                   use_container_width=True)

# ── Session State Init ─────────────────────────────────────────────────────────
if "holiday_list" not in st.session_state:
    st.session_state.holiday_list = []
if "ai_activities" not in st.session_state:        # confirmed / saved by user
    st.session_state.ai_activities = None
if "ai_activities_draft" not in st.session_state:  # raw from LLM, before save
    st.session_state.ai_activities_draft = None
if "ai_last_inputs" not in st.session_state:       # inputs used for last generation
    st.session_state.ai_last_inputs = {}

# ── Project Inputs ─────────────────────────────────────────────────────────────
st.header("1. Project Details")

project_name = st.text_input("Project Name", value="My Project")

col_l, col_r = st.columns(2)
with col_l:
    project_type = st.selectbox("Project Type", ["Residential", "Commercial", "Hospital"], index=1)
    floors = st.number_input("Number of Floors", min_value=1, value=5)
    built_up_area = st.number_input("Built-up Area (sqm)", min_value=100, value=1000)
with col_r:
    structural_system = st.selectbox("Structural System", ["Concrete", "Steel"])
    basements = st.number_input("Number of Basements", min_value=0, value=0)
    start_date = st.date_input("Project Start Date")

project_notes = st.text_area(
    "Additional Project Information",
    placeholder=(
        "Describe special requirements, phasing constraints, preferred activity breakdown, "
        "site conditions, client preferences, etc. The AI will use this to tailor activities."
    ),
    height=100,
)

# ── Calendar ───────────────────────────────────────────────────────────────────
st.header("2. Calendar")

calendar_folder = "Calendars"
os.makedirs(calendar_folder, exist_ok=True)

protected_calendars = ["india_5day", "india_6day"]
all_calendars = sorted({f.replace(".json", "") for f in os.listdir(calendar_folder) if f.endswith(".json")})
custom_calendars = sorted([c for c in all_calendars if c.lower() not in protected_calendars])

calendar_list = list(all_calendars)
calendar_list.append("➕ Create New Calendar")
if custom_calendars:
    calendar_list.append("🗑 Delete Calendar")

calendar_option = st.selectbox("Select Calendar", calendar_list)

if calendar_option == "➕ Create New Calendar":
    new_calendar_name = st.text_input("Calendar Name", value="My Calendar")
    working_days_input = st.multiselect(
        "Working Days",
        ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
        default=["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
    )
    st.subheader("Holidays")
    holiday_date = st.date_input("Select Holiday", value=None, key="hol_date_input")
    ch1, ch2 = st.columns(2)
    with ch1:
        if st.button("Add Holiday", key="add_holiday"):
            if holiday_date:
                hs = holiday_date.strftime("%Y-%m-%d")
                if hs not in st.session_state.holiday_list:
                    st.session_state.holiday_list.append(hs)
                    st.rerun()
    with ch2:
        if st.button("Clear Holidays", key="clear_holidays"):
            st.session_state.holiday_list = []
            st.rerun()

if st.session_state.holiday_list:
    st.dataframe(pd.DataFrame({"Holiday": st.session_state.holiday_list}), use_container_width=True)
    holiday_to_delete = st.selectbox("Remove Holiday", [""] + st.session_state.holiday_list)
    if st.button("Delete Holiday", key="delete_holiday"):
        if holiday_to_delete:
            st.session_state.holiday_list.remove(holiday_to_delete)
            st.rerun()

if calendar_option == "🗑 Delete Calendar":
    if not custom_calendars:
        st.info("No custom calendars available.")
    else:
        cal_del = st.selectbox("Select Calendar To Delete", custom_calendars, key="calendar_to_delete")
        if st.button("Delete Selected Calendar", key="delete_calendar"):
            os.remove(os.path.join(calendar_folder, f"{cal_del}.json"))
            st.success(f"{cal_del} deleted.")
            st.rerun()

if calendar_option == "➕ Create New Calendar":
    if st.button("Save New Calendar", key="save_calendar"):
        cal_data = {
            "CalendarName": new_calendar_name,
            "WorkingDays": working_days_input,
            "Holidays": st.session_state.holiday_list,
        }
        with open(os.path.join(calendar_folder, f"{new_calendar_name}.json"), "w") as fh:
            json.dump(cal_data, fh, indent=4)
        st.success("Calendar Saved")
        st.session_state.holiday_list = []
        st.rerun()

# ── Scheduling Rules Editor ────────────────────────────────────────────────────
st.header("3. Scheduling Logic Rules")

_rules_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "scheduling_rules.md")

with st.expander("📋 View / Edit Scheduling Rules", expanded=False):
    st.caption(
        "These rules are injected into every AI generation call. "
        "Edit bullet points to refine the output. Lines starting with `#` are comments (ignored by AI). "
        "Save after editing."
    )
    if os.path.exists(_rules_path):
        with open(_rules_path, "r", encoding="utf-8") as _rf:
            _current_rules = _rf.read()
    else:
        _current_rules = ""

    _edited_rules = st.text_area(
        "Scheduling Rules (Markdown)",
        value=_current_rules,
        height=400,
        key="rules_editor",
        label_visibility="collapsed",
    )
    if st.button("💾 Save Rules", key="save_rules"):
        with open(_rules_path, "w", encoding="utf-8") as _wf:
            _wf.write(_edited_rules)
        st.success("Rules saved. They will be used in the next AI generation.")

# ── AI Activity Generation ─────────────────────────────────────────────────────
st.header("4. Generate Activities")

col_ai1, col_ai2, col_ai3 = st.columns([2, 2, 3])
with col_ai1:
    if st.button("🤖 Generate Activities (AI)", key="gen_ai"):
        with st.spinner("Calling AI — this may take 15–30 seconds…"):
            try:
                user_inputs = {
                    "project_name": project_name,
                    "project_type": project_type,
                    "structural_system": structural_system,
                    "floors": int(floors),
                    "basements": int(basements),
                    "built_up_area": float(built_up_area),
                    "notes": project_notes,
                }
                acts = generate_activities(user_inputs, max_activities=100)
                # Keep schedule fields + Notes (rationale)
                draft = [
                    {k: v for k, v in a.items() if k in ("ID", "Activity", "Duration", "Predecessor", "WBS", "Notes")}
                    for a in acts
                ]
                st.session_state.ai_activities_draft = draft
                st.session_state.ai_last_inputs = user_inputs   # for rationale tab
                st.session_state.ai_activities = None   # clear previous saved set
                st.success(f"AI generated {len(draft)} activities. Review and save below.")
                st.rerun()
            except Exception as e:
                st.error(f"AI generation failed: {e}")

with col_ai2:
    if st.button("🔑 Test AI Connection", key="test_ai"):
        try:
            st.info(test_connection())
        except Exception as e:
            st.error(str(e))

with col_ai3:
    if st.session_state.ai_activities is not None:
        st.success(f"✅ {len(st.session_state.ai_activities)} AI activities saved and ready.")
        if st.button("🗑 Clear AI Activities", key="clear_ai"):
            st.session_state.ai_activities = None
            st.rerun()

# ── Activity Review / Editor ───────────────────────────────────────────────────
if st.session_state.ai_activities_draft is not None:
    st.subheader("Review & Edit AI Activities")

    draft_df = pd.DataFrame(st.session_state.ai_activities_draft)
    for col in ["ID", "Activity", "Duration", "Predecessor", "WBS", "Notes"]:
        if col not in draft_df.columns:
            draft_df[col] = ""
    draft_df = draft_df[["ID", "Activity", "Duration", "Predecessor", "WBS", "Notes"]]

    # Normalize WBS to plain string (AI may return lists like ["Phase","Sub"])
    def _wbs_to_str(v):
        if isinstance(v, list):
            return " > ".join(str(x) for x in v)
        return str(v) if v is not None else ""
    draft_df["WBS"] = draft_df["WBS"].apply(_wbs_to_str)

    # Sort by WBS for hierarchical display (preserve original order within same WBS)
    draft_df_sorted = draft_df.sort_values("WBS", kind="stable").reset_index(drop=True)

    tab_view, tab_rationale, tab_edit = st.tabs(["📊 WBS Hierarchy View", "💡 AI Rationale", "✏️ Edit Activities"])

    with tab_view:
        st.caption("Activities grouped by WBS level. Grey rows are WBS group headers.")
        wbs_html = render_wbs_table(draft_df_sorted, ["ID", "Duration", "Predecessor"])
        st.markdown(wbs_html, unsafe_allow_html=True)

    with tab_rationale:
        # ── Parameter summary card ─────────────────────────────────────────
        st.caption("Verify the parameters used to generate these activities, then review the AI's reasoning per activity.")
        params_used = st.session_state.get("ai_last_inputs", {})
        if params_used:
            pc1, pc2, pc3 = st.columns(3)
            with pc1:
                st.metric("Project Type", params_used.get("project_type", "—"))
                st.metric("Structural System", params_used.get("structural_system", "—"))
            with pc2:
                st.metric("Floors", params_used.get("floors", "—"))
                st.metric("Basements", params_used.get("basements", "—"))
            with pc3:
                st.metric("Built-up Area (sqm)", f"{float(params_used.get('built_up_area', 0)):,.0f}")
                st.metric("Activities Generated", len(draft_df_sorted))
            if params_used.get("notes", "").strip():
                st.info(f"**Special requirements passed to AI:** {params_used['notes'].strip()}")
            st.divider()
        # ── Per-activity rationale ─────────────────────────────────────────
        has_notes = draft_df_sorted["Notes"].astype(str).str.strip().replace("nan", "").replace("None", "")
        if has_notes.eq("").all():
            st.info("No rationale was returned by the AI. Re-generate to get notes.")
        else:
            for _, row in draft_df_sorted.iterrows():
                note = str(row.get("Notes", "") or "").strip()
                if note in ("", "nan", "None"):
                    continue
                wbs_label = str(row.get("WBS", "") or "")
                act_label = f"**{row['ID']}** — {row['Activity']}"
                duration_label = f"`{int(row['Duration'])} days`" if str(row.get("Duration", "")) not in ("", "nan") else ""
                with st.expander(f"{act_label}  {duration_label}", expanded=False):
                    if wbs_label and wbs_label not in ("nan", "None"):
                        st.caption(f"WBS: {wbs_label}")
                    st.write(note)

    with tab_edit:
        st.caption(
            "Edit any cell, add or delete rows. "
            "Predecessor: comma-separated IDs (e.g. `A1,A2`). "
            "WBS levels separated by ` > ` (e.g. `3. Superstructure > Floor 1`)."
        )
        edited_df = st.data_editor(
            draft_df,
            num_rows="dynamic",
            use_container_width=True,
            key="activity_editor",
            column_config={
                "ID":          st.column_config.TextColumn("ID", width="small"),
                "Activity":    st.column_config.TextColumn("Activity Name", width="large"),
                "Duration":    st.column_config.NumberColumn("Duration (days)", min_value=1, width="small"),
                "Predecessor": st.column_config.TextColumn("Predecessor IDs", width="medium"),
                "WBS":         st.column_config.TextColumn("WBS (levels separated by  >)", width="large"),
                "Notes":       st.column_config.TextColumn("AI Rationale / Notes", width="large"),
            },
        )

        cs1, cs2 = st.columns([1, 1])
        with cs1:
            if st.button("✅ Save Activities", key="save_activities"):
                saved = edited_df.dropna(subset=["ID", "Activity"]).to_dict(orient="records")
                st.session_state.ai_activities = saved
                st.session_state.ai_activities_draft = None
                st.success(f"Saved {len(saved)} activities. Select a calendar and click Generate Schedule.")
                st.rerun()
        with cs2:
            if st.button("🗑 Discard", key="discard_activities"):
                st.session_state.ai_activities_draft = None
                st.rerun()

# ── Generate Schedule ──────────────────────────────────────────────────────────
st.header("5. Generate Schedule")

if st.session_state.ai_activities_draft is not None:
    st.info("You have unsaved AI activities. Visit **✏️ Edit Activities** above to save or discard them before generating the schedule.")
else:
    if st.button("▶ Generate Schedule", key="generate_schedule"):
        if calendar_option in ["➕ Create New Calendar", "🗑 Delete Calendar"]:
            st.error("Please select a valid calendar before generating the schedule.")
            st.stop()

        # ── Load Calendar ──────────────────────────────────────────────────────
        cal_path = os.path.join(calendar_folder, f"{calendar_option}.json")
        with open(cal_path, "r") as fh:
            cal = json.load(fh)
        working_days = cal["WorkingDays"]
        holidays = cal.get("Holidays", [])

        # ── Resolve Activities Source ──────────────────────────────────────────
        using_ai = st.session_state.ai_activities is not None
        if using_ai:
            activities = st.session_state.ai_activities
        else:
            # Fall back to template file
            def find_template(ptype, ssys):
                roots = [d for d in os.listdir(".") if os.path.isdir(d) and d.lower() == "templates"]
                if not roots:
                    roots = [p for p in ["templates", "Templates"] if os.path.isdir(p)]
                if not roots:
                    roots = ["templates"]
                for root in roots:
                    for v in [ptype, ptype.lower(), ptype.capitalize()]:
                        c = os.path.join(root, v, f"{ssys}.json")
                        if os.path.exists(c):
                            return c
                for root in roots:
                    for dp, _, fnames in os.walk(root):
                        for fn in fnames:
                            if fn.lower() == f"{ssys.lower()}.json":
                                return os.path.join(dp, fn)
                return os.path.join("templates", ptype, f"{ssys}.json")

            template_path = find_template(project_type, structural_system)
            if not os.path.exists(template_path):
                st.error(f"Template not found: {template_path}. Generate AI activities first or add a template file.")
                st.stop()
            with open(template_path, "r") as fh:
                activities = json.load(fh)

        # ── Build DataFrame ────────────────────────────────────────────────────
        df = pd.DataFrame(activities)
        df["ID"] = df["ID"].astype(str)
        df["Start Date"] = ""
        df["Finish Date"] = ""

        # Ensure WBS column exists
        if "WBS" not in df.columns:
            df["WBS"] = "General"

        # Normalize predecessors into a list of string IDs
        def normalize_pred(raw):
            if isinstance(raw, list):
                return [str(x).strip() for x in raw if str(x).strip()]
            if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                return []
            if isinstance(raw, (int, float)):
                return [str(int(raw))]
            s = str(raw).strip()
            if not s:
                return []
            if s.startswith("[") and s.endswith("]"):
                try:
                    parsed = json.loads(s)
                    if isinstance(parsed, list):
                        items = []
                        for p in parsed:
                            if isinstance(p, dict):
                                items.append(str(p.get("id", p.get("ID", ""))))
                            else:
                                items.append(str(p).strip())
                        return [x for x in items if x]
                except Exception:
                    pass
            if "," in s:
                return [p.strip() for p in s.split(",") if p.strip()]
            return [s]

        pred_col = "Predecessor" if "Predecessor" in df.columns else "Predecessors"
        df["Predecessors"] = df[pred_col].apply(normalize_pred)

        # ── Duration scaling (template mode only) ──────────────────────────────
        if not using_ai:
            floor_factor = max(1.0, int(floors) / 5)
            df["Duration"] = (df["Duration"] * floor_factor).round().astype(int)
            if int(basements) > 0:
                mask = df["Activity"].str.contains("Excavation", case=False, na=False)
                df.loc[mask, "Duration"] += int(basements) * 5
            if float(built_up_area) > 5000:
                df["Duration"] = (df["Duration"] * 1.2).round().astype(int)

        # ── Build topological order and compute dates ──────────────────────────
        id_map = {str(r["ID"]): i for i, r in df.iterrows()}

        # Validate all predecessors exist before proceeding
        errors = []
        for _, row in df.iterrows():
            for pid in row["Predecessors"]:
                if pid not in id_map:
                    errors.append(f"Activity '{row['Activity']}' (ID {row['ID']}) references unknown predecessor '{pid}'")
        if errors:
            for e in errors:
                st.error(e)
            st.stop()

        adj = defaultdict(list)
        indeg = {i: 0 for i in df.index}
        for i, row in df.iterrows():
            for pid in row["Predecessors"]:
                src = id_map[pid]
                adj[src].append(i)
                indeg[i] += 1

        q = deque(n for n, d in indeg.items() if d == 0)
        topo = []
        while q:
            n = q.popleft()
            topo.append(n)
            for s in adj[n]:
                indeg[s] -= 1
                if indeg[s] == 0:
                    q.append(s)

        if len(topo) != len(df):
            st.error("Cycle detected in predecessor relationships. Please review your activity list.")
            st.stop()

        project_start = pd.to_datetime(start_date)
        finish_dates = {}   # str(ID) -> finish datetime

        for i in topo:
            row = df.loc[i]
            preds = row["Predecessors"]
            if not preds:
                act_start = project_start
            else:
                latest = max(finish_dates[pid] for pid in preds)
                act_start = pd.to_datetime(next_working_day(latest, working_days, holidays))
            act_finish = add_working_days(act_start, int(row["Duration"]), working_days, holidays)
            df.at[i, "Start Date"] = act_start.strftime("%d-%b-%Y")
            df.at[i, "Finish Date"] = act_finish.strftime("%d-%b-%Y")
            finish_dates[str(row["ID"])] = act_finish

        # ── CPM ───────────────────────────────────────────────────────────────
        try:
            df = calculate_cpm(df)
        except Exception as cpm_err:
            st.warning(f"CPM calculation skipped: {cpm_err}")
            for col in ["Float", "Critical", "ES", "EF", "LS", "LF"]:
                if col not in df.columns:
                    df[col] = ""

        # ── Build display predecessor / relationship / lag columns ─────────────
        def pred_display(plist):
            ids, rels, lags = [], [], []
            items = plist if isinstance(plist, list) else [plist]
            for item in items:
                if item is None:
                    continue
                if isinstance(item, dict):
                    ids.append(str(item.get("id", "")))
                    rels.append(str(item.get("type", "FS")).upper())
                    lags.append(str(item.get("lag", 0)))
                elif isinstance(item, str):
                    parts = item.split(":")
                    ids.append(parts[0])
                    rels.append(parts[1].upper() if len(parts) > 1 else "FS")
                    lags.append(parts[2] if len(parts) > 2 else "0")
                else:
                    ids.append(str(int(item)) if isinstance(item, (int, float)) else str(item))
                    rels.append("FS")
                    lags.append("0")
            return pd.Series({
                "predecessor": ",".join(ids),
                "Relationship": ",".join(rels),
                "lag": ",".join(lags),
            })

        df = pd.concat([df, df["Predecessors"].apply(pred_display)], axis=1)

        # ── Display ────────────────────────────────────────────────────────────
        st.subheader("Project Details")
        details = {
            "Project Name": project_name,
            "Project Type": project_type,
            "Structural System": structural_system,
            "Floors": int(floors),
            "Basements": int(basements),
            "Built-up Area (sqm)": float(built_up_area),
            "Start Date": str(start_date),
            "Activities Source": "AI Generated" if using_ai else "Template",
        }
        st.table(pd.DataFrame(details.items(), columns=["Field", "Value"]))

        st.subheader("Calendar Details")
        st.write("Calendar:", calendar_option)
        st.write("Working Days:", ", ".join(working_days))
        if holidays:
            st.write("Holidays:", ", ".join(holidays))

        st.subheader("Generated Schedule")
        sched_data_cols = ["ID", "Duration", "Start Date", "Finish Date",
                           "predecessor", "Relationship", "lag", "Float", "Critical"]
        for c in sched_data_cols + ["WBS", "Activity"]:
            if c not in df.columns:
                df[c] = ""
        # Sort by WBS for hierarchical display, then by Start Date within each WBS group
        df_sched = df.copy()
        def _wbs_to_str_sched(v):
            if isinstance(v, list):
                return " > ".join(str(x) for x in v)
            return str(v) if v is not None else ""
        df_sched["WBS"] = df_sched["WBS"].apply(_wbs_to_str_sched)
        df_sched["_sort_wbs"] = df_sched["WBS"]
        try:
            df_sched["_sort_date"] = pd.to_datetime(df_sched["Start Date"], format="%d-%b-%Y", errors="coerce")
        except Exception:
            df_sched["_sort_date"] = pd.NaT
        df_sched = df_sched.sort_values(["_sort_wbs", "_sort_date"], kind="stable").reset_index(drop=True)
        df_sched = df_sched.drop(columns=["_sort_wbs", "_sort_date"])

        # ── Inject Milestones WBS (always first) ───────────────────────────────
        def _phase_finish(frame, *keywords):
            """Latest Finish Date among rows whose WBS contains any of the keywords."""
            pattern = "|".join(keywords)
            mask = frame["WBS"].str.contains(pattern, case=False, na=False, regex=True)
            if not mask.any():
                return None
            dates = pd.to_datetime(frame.loc[mask, "Finish Date"], format="%d-%b-%Y", errors="coerce")
            return None if dates.isna().all() else dates.max().strftime("%d-%b-%Y")

        _all_starts   = pd.to_datetime(df_sched["Start Date"],  format="%d-%b-%Y", errors="coerce")
        _all_finishes = pd.to_datetime(df_sched["Finish Date"], format="%d-%b-%Y", errors="coerce")
        _proj_start_ms  = _all_starts.min().strftime("%d-%b-%Y")   if not _all_starts.isna().all()   else None
        _proj_finish_ms = _all_finishes.max().strftime("%d-%b-%Y") if not _all_finishes.isna().all() else None

        # (id, name, date, kind)  kind="start" → show Start Date only; "finish" → show Finish Date only
        _ms_defs = [
            ("MS01", "Project Start",           _proj_start_ms,  "start"),
            ("MS02", "Mobilization Complete",
             _phase_finish(df_sched, r"mobil"),                  "finish"),
            ("MS03", "Substructure Complete",
             _phase_finish(df_sched, r"substructure", r"basement", r"foundati", r"sub.?str"), "finish"),
            ("MS04", "Superstructure Complete",
             _phase_finish(df_sched, r"superstructure", r"super.?str", r"structural frame", r"rcc frame"), "finish"),
            ("MS05", "Finishes Complete",
             _phase_finish(df_sched, r"finish", r"interior", r"facade", r"cladding"), "finish"),
            ("MS06", "Project Completion",      _proj_finish_ms, "finish"),
        ]

        _ms_rows = []
        for _ms_id, _ms_name, _ms_date, _ms_kind in _ms_defs:
            if not _ms_date:
                continue
            _ms_rows.append({
                "ID": _ms_id, "Activity": _ms_name,
                "Duration": 0,
                "Start Date":  _ms_date if _ms_kind == "start" else "",
                "Finish Date": _ms_date if _ms_kind == "finish" else "",
                "WBS": "Milestones", "predecessor": "", "Relationship": "",
                "lag": "", "Float": 0, "Critical": "True",
            })

        if _ms_rows:
            _ms_df = pd.DataFrame(_ms_rows).reindex(columns=df_sched.columns, fill_value="")
            df_sched = pd.concat([_ms_df, df_sched], ignore_index=True)

        sched_html = render_wbs_table(df_sched, sched_data_cols)
        st.markdown(sched_html, unsafe_allow_html=True)

        # ── Styled Excel export (mirrors WBS HTML table) ──────────────────────
        import io
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        def _build_excel_wbs(source_df, excel_data_cols):
            """Build an openpyxl Workbook that mirrors the WBS hierarchical table."""
            def _pw(v):
                s = str(v).strip() if v is not None else ""
                if s in ("", "nan", "None", "General"):
                    return ["General"]
                parts = [p.strip() for p in re.split(r"\s*>\s*", s) if p.strip()]
                return parts if parts else ["General"]

            # --- Style definitions ---
            # WBS header level fills (dark → light grey, matching HTML table)
            LVL_FG   = ["455A64", "607D8B", "90A4AE", "B0BEC5", "CFD8DC"]
            LVL_FONT = ["FFFFFF", "FFFFFF", "263238", "263238", "263238"]
            HDR_FILL = PatternFill("solid", fgColor="263238")
            HDR_FONT = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
            CRIT_FILL = PatternFill("solid", fgColor="FFF8E1")
            CRIT_FONT = Font(color="C62828", bold=True, size=10, name="Calibri")
            NORM_FONT = Font(size=10, name="Calibri")
            ALT_FILLS = [
                PatternFill("solid", fgColor="FFFFFF"),
                PatternFill("solid", fgColor="F5F7F9"),
            ]
            thin = None  # no border for cleaner look

            wb = Workbook()
            ws = wb.active
            ws.title = "Schedule"

            # Column layout: Activity first, then data cols
            all_cols = ["Activity"] + excel_data_cols
            col_widths = {
                "Activity": 42, "ID": 12, "Duration": 12,
                "Start Date": 14, "Finish Date": 14,
                "predecessor": 22, "Relationship": 14,
                "lag": 8, "Float": 8, "Critical": 10,
            }

            # Header row
            ws.append(all_cols)
            for cell in ws[1]:
                cell.fill = HDR_FILL
                cell.font = HDR_FONT
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
            ws.row_dimensions[1].height = 22

            current_path = []
            excel_row = 1
            alt_idx = 0  # alternating row counter for activity rows only

            for _, row in source_df.iterrows():
                wbs_path = _pw(row.get("WBS", ""))
                common = sum(1 for a, b in zip(current_path, wbs_path) if a == b)

                # WBS header rows for new levels
                for lvl in range(common, len(wbs_path)):
                    excel_row += 1
                    fg   = LVL_FG[min(lvl, len(LVL_FG) - 1)]
                    fc   = LVL_FONT[min(lvl, len(LVL_FONT) - 1)]
                    fill = PatternFill("solid", fgColor=fg)
                    font = Font(color=fc, bold=True, size=10, name="Calibri")
                    label = "  " * (lvl * 2) + wbs_path[lvl]
                    ws.append([""] * len(all_cols))
                    ws.cell(excel_row, 1).value = label
                    for c in range(1, len(all_cols) + 1):
                        cell = ws.cell(excel_row, c)
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[excel_row].height = 18

                current_path = list(wbs_path)

                # Activity row
                excel_row += 1
                depth = len(wbs_path)
                is_crit = str(row.get("Critical", "")).strip().lower() in ("true", "1", "yes")
                act_val = row.get("Activity", "")

                row_data = [act_val] + [
                    (row.get(col, "") if row.get(col, "") not in (None, "nan") else "")
                    for col in excel_data_cols
                ]
                ws.append(row_data)

                row_fill = CRIT_FILL if is_crit else ALT_FILLS[alt_idx % 2]
                row_font = CRIT_FONT if is_crit else NORM_FONT
                for c in range(1, len(all_cols) + 1):
                    cell = ws.cell(excel_row, c)
                    cell.fill = row_fill
                    cell.font = row_font
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                # Indent activity name cell
                ws.cell(excel_row, 1).alignment = Alignment(
                    horizontal="left", vertical="center", indent=depth
                )
                ws.row_dimensions[excel_row].height = 16
                alt_idx += 1

            # Column widths
            for i, col_name in enumerate(all_cols, 1):
                ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 14)

            # Freeze header row
            ws.freeze_panes = "A2"
            return wb

        excel_data_cols = ["ID", "Duration", "Start Date", "Finish Date",
                           "predecessor", "Relationship", "lag", "Float", "Critical"]
        for c in excel_data_cols + ["WBS", "Activity"]:
            if c not in df_sched.columns:
                df_sched[c] = ""

        wb = _build_excel_wbs(df_sched, excel_data_cols)
        excel_buf = io.BytesIO()
        wb.save(excel_buf)
        excel_buf.seek(0)

        def _build_p6_xml(source_df):
            """Build a Primavera P6 compatible XML string from the schedule dataframe."""
            import xml.etree.ElementTree as ET
            from xml.dom import minidom

            root_el = ET.Element("APIBusinessObjects")

            proj_oid = "1"
            proj_id = "PROJ001"
            proj_el = ET.SubElement(root_el, "Project")
            ET.SubElement(proj_el, "ObjectId").text = proj_oid
            ET.SubElement(proj_el, "Id").text = proj_id
            ET.SubElement(proj_el, "Name").text = "Schedule"

            wbs_map = {}
            wbs_oid_counter = [3]
            root_wbs_oid = "2"
            root_wbs_el = ET.SubElement(root_el, "WBS")
            ET.SubElement(root_wbs_el, "ObjectId").text = root_wbs_oid
            ET.SubElement(root_wbs_el, "ProjectObjectId").text = proj_oid
            ET.SubElement(root_wbs_el, "Code").text = proj_id
            ET.SubElement(root_wbs_el, "Name").text = "Schedule"
            ET.SubElement(root_wbs_el, "ParentObjectId").text = ""

            def _pw_tuple(v):
                s = str(v).strip() if v is not None else ""
                if s in ("", "nan", "None", "General"):
                    return ("General",)
                parts = tuple(p.strip() for p in re.split(r"\s*>\s*", s) if p.strip())
                return parts if parts else ("General",)

            def _get_wbs_oid(wbs_path_tuple):
                if wbs_path_tuple in wbs_map:
                    return wbs_map[wbs_path_tuple]
                parent_oid = root_wbs_oid if len(wbs_path_tuple) == 1 else _get_wbs_oid(wbs_path_tuple[:-1])
                oid = str(wbs_oid_counter[0])
                wbs_oid_counter[0] += 1
                wbs_map[wbs_path_tuple] = oid
                wbs_el = ET.SubElement(root_el, "WBS")
                ET.SubElement(wbs_el, "ObjectId").text = oid
                ET.SubElement(wbs_el, "ProjectObjectId").text = proj_oid
                ET.SubElement(wbs_el, "Code").text = ".".join(wbs_path_tuple)
                ET.SubElement(wbs_el, "Name").text = wbs_path_tuple[-1]
                ET.SubElement(wbs_el, "ParentObjectId").text = parent_oid
                return oid

            seen_wbs = []
            for _, row in source_df.iterrows():
                wbs_tuple = _pw_tuple(row.get("WBS", ""))
                for i in range(1, len(wbs_tuple) + 1):
                    sub = wbs_tuple[:i]
                    if sub not in seen_wbs:
                        seen_wbs.append(sub)
                        _get_wbs_oid(sub)

            def _fmt_date(d_str):
                if not d_str or str(d_str).strip() in ("", "nan", "None"):
                    return ""
                try:
                    dt = pd.to_datetime(str(d_str).strip(), format="%d-%b-%Y", errors="coerce")
                    return "" if pd.isna(dt) else dt.strftime("%Y-%m-%dT08:00:00")
                except Exception:
                    return ""

            def _rel_type(r_str):
                mapping = {"FS": "FinishToStart", "SS": "StartToStart",
                           "FF": "FinishToFinish", "SF": "StartToFinish"}
                return mapping.get(str(r_str).strip().upper(), "FinishToStart")

            def _hours(val):
                try:
                    return float(str(val).strip()) * 8 if str(val).strip() not in ("", "nan", "None") else 0.0
                except (ValueError, TypeError):
                    return 0.0

            act_oid_map = {}
            act_oid_counter = [1000]
            rel_oid_counter = [5000]
            relationships = []

            for _, row in source_df.iterrows():
                act_id_raw = str(row.get("ID", "")).strip()
                if not act_id_raw or act_id_raw in ("", "nan", "None"):
                    continue
                oid = str(act_oid_counter[0])
                act_oid_counter[0] += 1
                act_oid_map[act_id_raw] = oid

                wbs_tuple = _pw_tuple(row.get("WBS", ""))
                wbs_oid = wbs_map.get(wbs_tuple, root_wbs_oid)
                dur_h = _hours(row.get("Duration", 0))
                float_h = _hours(row.get("Float", 0))
                is_crit = str(row.get("Critical", "")).strip().lower() in ("true", "1", "yes")

                act_el = ET.SubElement(root_el, "Activity")
                ET.SubElement(act_el, "ObjectId").text = oid
                ET.SubElement(act_el, "ProjectObjectId").text = proj_oid
                ET.SubElement(act_el, "WBSObjectId").text = wbs_oid
                ET.SubElement(act_el, "Id").text = act_id_raw
                ET.SubElement(act_el, "Name").text = str(row.get("Activity", "")).strip()
                ET.SubElement(act_el, "Type").text = "TaskDependent"
                ET.SubElement(act_el, "Status").text = "Not Started"
                ET.SubElement(act_el, "PlannedDuration").text = str(dur_h)
                ET.SubElement(act_el, "RemainingDuration").text = str(dur_h)
                ET.SubElement(act_el, "PlannedStartDate").text = _fmt_date(row.get("Start Date", ""))
                ET.SubElement(act_el, "PlannedFinishDate").text = _fmt_date(row.get("Finish Date", ""))
                ET.SubElement(act_el, "StartDate").text = _fmt_date(row.get("Start Date", ""))
                ET.SubElement(act_el, "FinishDate").text = _fmt_date(row.get("Finish Date", ""))
                ET.SubElement(act_el, "TotalFloat").text = str(float_h)
                ET.SubElement(act_el, "CriticalFlag").text = "true" if is_crit else "false"
                ET.SubElement(act_el, "DurationPercentComplete").text = "0"
                ET.SubElement(act_el, "ActualDuration").text = "0"

                pred_str = str(row.get("predecessor", "")).strip()
                if pred_str and pred_str not in ("", "nan", "None"):
                    for pred_id in re.split(r"[,;]+", pred_str):
                        pred_id = pred_id.strip()
                        if pred_id:
                            relationships.append({
                                "pred_id": pred_id,
                                "succ_id": act_id_raw,
                                "rel_type": _rel_type(row.get("Relationship", "FS")),
                                "lag": _hours(row.get("lag", 0)),
                            })

            for rel in relationships:
                pred_oid = act_oid_map.get(rel["pred_id"])
                succ_oid = act_oid_map.get(rel["succ_id"])
                if pred_oid and succ_oid:
                    rel_el = ET.SubElement(root_el, "Relationship")
                    ET.SubElement(rel_el, "ObjectId").text = str(rel_oid_counter[0])
                    rel_oid_counter[0] += 1
                    ET.SubElement(rel_el, "ProjectObjectId").text = proj_oid
                    ET.SubElement(rel_el, "PredecessorActivityObjectId").text = pred_oid
                    ET.SubElement(rel_el, "SuccessorActivityObjectId").text = succ_oid
                    ET.SubElement(rel_el, "Type").text = rel["rel_type"]
                    ET.SubElement(rel_el, "Lag").text = str(rel["lag"])

            xml_str = ET.tostring(root_el, encoding="unicode")
            dom = minidom.parseString(xml_str)
            return dom.toprettyxml(indent="  ", encoding=None)

        p6_xml_str = _build_p6_xml(df_sched)
        p6_buf = io.BytesIO(p6_xml_str.encode("utf-8"))
        p6_buf.seek(0)

        def _build_quality_log(source_df, wd, hols, thresh_dur=20, thresh_float=20):
            """Run schedule quality checks and return an openpyxl Workbook."""
            issues = []

            # Build successor map from the flattened predecessor column
            succ_map = defaultdict(set)
            for _, row in source_df.iterrows():
                pred_str = str(row.get("predecessor", "")).strip()
                if pred_str and pred_str not in ("nan", "None"):
                    for pid in re.split(r"[,;]+", pred_str):
                        pid = pid.strip()
                        if pid:
                            succ_map[pid].add(str(row["ID"]))

            # Build predecessor set per activity
            pred_map_q = {}
            for _, row in source_df.iterrows():
                pred_str = str(row.get("predecessor", "")).strip()
                preds = set()
                if pred_str and pred_str not in ("nan", "None"):
                    for pid in re.split(r"[,;]+", pred_str):
                        pid = pid.strip()
                        if pid:
                            preds.add(pid)
                pred_map_q[str(row["ID"])] = preds

            def _is_milestone(row):
                try:
                    return int(float(str(row.get("Duration", 1)))) == 0
                except Exception:
                    return False

            # ── 1. Open Starts ─────────────────────────────────────────────
            for _, row in source_df.iterrows():
                act_id = str(row["ID"])
                if not pred_map_q.get(act_id) and not _is_milestone(row):
                    issues.append({
                        "Check": "Open Start", "Severity": "Warning",
                        "ID": act_id, "Activity": str(row.get("Activity", "")),
                        "WBS": str(row.get("WBS", "")),
                        "Details": "No predecessors assigned (open start)",
                    })

            # ── 2. Open Ends ───────────────────────────────────────────────
            for _, row in source_df.iterrows():
                act_id = str(row["ID"])
                if act_id not in succ_map and not _is_milestone(row):
                    issues.append({
                        "Check": "Open End", "Severity": "Warning",
                        "ID": act_id, "Activity": str(row.get("Activity", "")),
                        "WBS": str(row.get("WBS", "")),
                        "Details": "No successors assigned (open end)",
                    })

            # ── 3. Excessive Duration ──────────────────────────────────────
            for _, row in source_df.iterrows():
                try:
                    dur = int(float(str(row.get("Duration", 0))))
                    if dur > thresh_dur and not _is_milestone(row):
                        issues.append({
                            "Check": "Excessive Duration", "Severity": "Warning",
                            "ID": str(row["ID"]), "Activity": str(row.get("Activity", "")),
                            "WBS": str(row.get("WBS", "")),
                            "Details": f"Duration = {dur} working days (threshold: {thresh_dur})",
                        })
                except Exception:
                    pass

            # ── 4. Missing Milestones ──────────────────────────────────────
            milestones_q = [row for _, row in source_df.iterrows() if _is_milestone(row)]
            if not milestones_q:
                issues.append({
                    "Check": "Missing Milestones", "Severity": "Warning",
                    "ID": "", "Activity": "", "WBS": "",
                    "Details": "No milestones (zero-duration activities) found in the schedule",
                })
            else:
                if not [r for r in milestones_q if not pred_map_q.get(str(r["ID"]))]:
                    issues.append({
                        "Check": "Missing Milestones", "Severity": "Warning",
                        "ID": "", "Activity": "", "WBS": "",
                        "Details": "No project start milestone (zero-duration, no predecessors) found",
                    })
                if not [r for r in milestones_q if str(r["ID"]) not in succ_map]:
                    issues.append({
                        "Check": "Missing Milestones", "Severity": "Warning",
                        "ID": "", "Activity": "", "WBS": "",
                        "Details": "No project finish milestone (zero-duration, no successors) found",
                    })

            # ── 5. Excessive Total Float ───────────────────────────────────
            for _, row in source_df.iterrows():
                flt_raw = str(row.get("Float", "")).strip()
                if flt_raw in ("", "nan", "None"):
                    continue
                try:
                    flt = int(float(flt_raw))
                    if flt > thresh_float:
                        issues.append({
                            "Check": "Excessive Float", "Severity": "Info",
                            "ID": str(row["ID"]), "Activity": str(row.get("Activity", "")),
                            "WBS": str(row.get("WBS", "")),
                            "Details": f"Total Float = {flt} working days (threshold: {thresh_float})",
                        })
                except Exception:
                    pass

            # ── 6. Circular Dependencies ───────────────────────────────────
            # A completed schedule means the topological sort succeeded → no cycle.
            issues.append({
                "Check": "Circular Dependencies", "Severity": "Pass",
                "ID": "", "Activity": "", "WBS": "",
                "Details": "No circular dependencies detected — schedule generated successfully",
            })

            # ── 7. WBS Completeness ────────────────────────────────────────
            for _, row in source_df.iterrows():
                wbs = str(row.get("WBS", "")).strip()
                if wbs in ("", "nan", "None", "General"):
                    issues.append({
                        "Check": "WBS Completeness", "Severity": "Info",
                        "ID": str(row["ID"]), "Activity": str(row.get("Activity", "")),
                        "WBS": wbs,
                        "Details": "Activity uses generic or missing WBS code",
                    })

            # ── 8. Calendar Inconsistencies ────────────────────────────────
            day_map_q = {
                "Monday": 0, "Tuesday": 1, "Wednesday": 2, "Thursday": 3,
                "Friday": 4, "Saturday": 5, "Sunday": 6,
            }
            wd_nums = {day_map_q[d] for d in wd if d in day_map_q}
            hol_set = set(hols)
            for _, row in source_df.iterrows():
                for date_col in ["Start Date", "Finish Date"]:
                    d_str = str(row.get(date_col, "")).strip()
                    if d_str in ("", "nan", "None"):
                        continue
                    try:
                        dt = pd.to_datetime(d_str, format="%d-%b-%Y", errors="coerce")
                        if pd.isna(dt):
                            continue
                        if dt.weekday() not in wd_nums:
                            issues.append({
                                "Check": "Calendar Inconsistency", "Severity": "Error",
                                "ID": str(row["ID"]), "Activity": str(row.get("Activity", "")),
                                "WBS": str(row.get("WBS", "")),
                                "Details": f"{date_col} {d_str} falls on a non-working day ({dt.strftime('%A')})",
                            })
                        if dt.strftime("%Y-%m-%d") in hol_set:
                            issues.append({
                                "Check": "Calendar Inconsistency", "Severity": "Error",
                                "ID": str(row["ID"]), "Activity": str(row.get("Activity", "")),
                                "WBS": str(row.get("WBS", "")),
                                "Details": f"{date_col} {d_str} falls on a calendar holiday",
                            })
                    except Exception:
                        pass

            # ── 9. Resource Assignments (future enhancement) ───────────────
            issues.append({
                "Check": "Resource Assignments", "Severity": "Info",
                "ID": "", "Activity": "", "WBS": "",
                "Details": "Resource assignment checks are reserved for a future enhancement",
            })

            # ── Build Excel workbook ───────────────────────────────────────
            wb_q = Workbook()

            SEV_COLORS = {
                "Error":   ("FFCDD2", "C62828"),
                "Warning": ("FFF9C4", "F57F17"),
                "Info":    ("E3F2FD", "1565C0"),
                "Pass":    ("E8F5E9", "2E7D32"),
            }
            HDR_FILL_Q = PatternFill("solid", fgColor="263238")
            HDR_FONT_Q = Font(color="FFFFFF", bold=True, size=10, name="Calibri")

            def _q_hdr(ws, cols):
                ws.append(cols)
                for cell in ws[1]:
                    cell.fill = HDR_FILL_Q
                    cell.font = HDR_FONT_Q
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.row_dimensions[1].height = 20

            # Summary sheet
            ws_sum = wb_q.active
            ws_sum.title = "Summary"
            _q_hdr(ws_sum, ["Check", "Issues", "Errors", "Warnings", "Info", "Status"])

            checks_order = [
                "Open Start", "Open End", "Excessive Duration", "Missing Milestones",
                "Excessive Float", "Circular Dependencies", "WBS Completeness",
                "Calendar Inconsistency", "Resource Assignments",
            ]
            for chk in checks_order:
                chk_iss = [i for i in issues if i["Check"] == chk]
                n_e = sum(1 for i in chk_iss if i["Severity"] == "Error")
                n_w = sum(1 for i in chk_iss if i["Severity"] == "Warning")
                n_i = sum(1 for i in chk_iss if i["Severity"] == "Info")
                n_p = sum(1 for i in chk_iss if i["Severity"] == "Pass")
                if n_p > 0 and n_e == 0 and n_w == 0:
                    status, sev, total = "PASS", "Pass", 0
                elif n_e > 0:
                    status, sev, total = "FAIL", "Error", n_e + n_w + n_i
                elif n_w > 0:
                    status, sev, total = "REVIEW", "Warning", n_w + n_i
                elif n_i > 0:
                    status, sev, total = "INFO", "Info", n_i
                else:
                    status, sev, total = "PASS", "Pass", 0

                ws_sum.append([chk, total, n_e, n_w, n_i, status])
                r = ws_sum.max_row
                bg, fc = SEV_COLORS[sev]
                for c in range(1, 7):
                    cell = ws_sum.cell(r, c)
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(color=fc, size=10, name="Calibri", bold=(c == 6))
                    cell.alignment = Alignment(
                        horizontal="left" if c == 1 else "center", vertical="center"
                    )
                ws_sum.row_dimensions[r].height = 16

            for col_ltr, width in [("A", 32), ("B", 10), ("C", 10), ("D", 12), ("E", 8), ("F", 10)]:
                ws_sum.column_dimensions[col_ltr].width = width
            ws_sum.freeze_panes = "A2"

            # Issues sheet
            ws_iss = wb_q.create_sheet("Issues")
            _q_hdr(ws_iss, ["Check", "Severity", "ID", "Activity", "WBS", "Details"])

            sev_order_map = {"Error": 0, "Warning": 1, "Info": 2}
            issues_only = sorted(
                [i for i in issues if i["Severity"] != "Pass"],
                key=lambda x: sev_order_map.get(x["Severity"], 3),
            )
            if not issues_only:
                ws_iss.append(["(No issues found)", "", "", "", "", "All checks passed"])
                ws_iss.cell(2, 1).font = Font(color="2E7D32", size=10, name="Calibri")
            else:
                for iss in issues_only:
                    ws_iss.append([
                        iss["Check"], iss["Severity"], iss["ID"],
                        iss["Activity"], iss["WBS"], iss["Details"],
                    ])
                    r = ws_iss.max_row
                    bg, fc = SEV_COLORS.get(iss["Severity"], ("FFFFFF", "000000"))
                    for c in range(1, 7):
                        cell = ws_iss.cell(r, c)
                        cell.fill = PatternFill("solid", fgColor=bg)
                        cell.font = Font(color=fc, size=10, name="Calibri")
                        cell.alignment = Alignment(
                            horizontal="left", vertical="center", wrap_text=(c == 6)
                        )
                    ws_iss.row_dimensions[r].height = 16

            for col_ltr, width in [("A", 28), ("B", 12), ("C", 10), ("D", 36), ("E", 30), ("F", 60)]:
                ws_iss.column_dimensions[col_ltr].width = width
            ws_iss.freeze_panes = "A2"

            return wb_q

        ql_wb = _build_quality_log(df_sched, working_days, holidays)
        ql_buf = io.BytesIO()
        ql_wb.save(ql_buf)
        ql_buf.seek(0)

        safe_name = re.sub(r"[^\w\-]", "_", project_name.strip()) or "project"
        dl_col1, dl_col2, dl_col3 = st.columns(3)
        with dl_col1:
            st.download_button(
                "📥 Download Excel Schedule",
                excel_buf,
                file_name="schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        with dl_col2:
            st.download_button(
                "📥 Download P6 XML",
                p6_buf,
                file_name="schedule_p6.xml",
                mime="application/xml",
            )
        with dl_col3:
            st.download_button(
                "📋 Download Quality Log",
                ql_buf,
                file_name=f"{safe_name}_quality_log.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
